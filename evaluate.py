
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuração ──────────────────────────────────────────────────────────
INPUT_FILE = r"C:\Users\112585631\Documents\Pessoal\MBA\Proj Final\Results\Métricas Quantitativas.xlsx"
OUTPUT_FILE = os.path.join(os.path.dirname(INPUT_FILE), "ACL_Evaluation_Report.xlsx")

NOT_FOUND_LABELS = {"Customer not found", "Não encontrado"}
HIGH_THRESHOLD = 90
MID_LOW = 60
MID_HIGH = 80


# ── Leitura dos dados ────────────────────────────────────────────────────
def load_data(filepath):
    model = pd.read_excel(filepath, sheet_name="Model Output")
    expected = pd.read_excel(filepath, sheet_name="Esperado")
    model.columns = model.columns.str.strip()
    expected.columns = expected.columns.str.strip()
    return model, expected


# ── Helpers ───────────────────────────────────────────────────────────────
def is_not_found(value):
    if pd.isna(value):
        return True
    return str(value).strip() in NOT_FOUND_LABELS


def get_summary_row(model_df, company):
    """Retorna a linha de maior probabilidade do modelo para uma empresa."""
    rows = model_df[model_df["Company"] == company]
    if rows.empty:
        return None
    valid = rows.dropna(subset=["Probability"])
    if valid.empty:
        return rows.iloc[0]
    return valid.loc[valid["Probability"].idxmax()]


def get_all_model_rows(model_df, company):
    return model_df[model_df["Company"] == company]


# ── Classificação por empresa ─────────────────────────────────────────────
def classify_companies(model_df, expected_df):
    companies = expected_df["Company"].unique()
    results = []

    for company in companies:
        exp_rows = expected_df[expected_df["Company"] == company]
        mod_rows = get_all_model_rows(model_df, company)
        best_model = get_summary_row(model_df, company)

        exp_not_found = all(is_not_found(r["ACCT LEVEL"]) for _, r in exp_rows.iterrows())
        mod_not_found = best_model is None or is_not_found(best_model.get("ACCT LEVEL"))

        best_prob = None if best_model is None else best_model.get("Probability")
        best_acct = None if best_model is None else best_model.get("ACCT")
        best_acct_name = None if best_model is None else best_model.get("ACCT NAME")
        best_level = None if best_model is None else best_model.get("ACCT LEVEL")

        exp_accts = set(exp_rows["ACCT"].dropna().astype(str))

        # Classificação
        if exp_not_found and mod_not_found:
            classification = "True Negative"
            detail = "Modelo e esperado concordam: cliente não encontrado."

        elif exp_not_found and not mod_not_found:
            classification = "False Positive"
            detail = (
                f"Modelo recomendou {best_acct} ({best_acct_name}) "
                f"com score {best_prob}, mas o esperado é 'Não encontrado'."
            )

        elif not exp_not_found and mod_not_found:
            classification = "False Negative"
            detail = "Modelo não encontrou o cliente, mas ele existe no esperado."

        elif not exp_not_found and not mod_not_found:
            if str(best_acct) in exp_accts:
                classification = "True Positive"
                detail = f"Match correto: {best_acct} ({best_acct_name}) com score {best_prob}."
            else:
                any_match = any(
                    str(r.get("ACCT")) in exp_accts
                    for _, r in mod_rows.iterrows()
                    if not is_not_found(r.get("ACCT LEVEL"))
                )
                if any_match:
                    classification = "Partial Match"
                    detail = (
                        f"O candidato de maior score ({best_acct}, score {best_prob}) "
                        f"não é o esperado, mas o esperado aparece entre os outros candidatos."
                    )
                else:
                    classification = "Wrong Match"
                    detail = (
                        f"Modelo recomendou {best_acct} ({best_acct_name}) "
                        f"com score {best_prob}, mas o esperado era diferente."
                    )
        else:
            classification = "Unclassified"
            detail = ""

        # Flags adicionais
        n_model_rows = len(mod_rows)
        has_multiple_100 = False
        if not mod_rows.empty:
            rows_100 = mod_rows[mod_rows["Probability"] == 100]
            unique_accts_100 = rows_100["ACCT"].nunique()
            has_multiple_100 = unique_accts_100 > 1

        prob_band = ""
        if best_prob is not None and not pd.isna(best_prob):
            if best_prob >= HIGH_THRESHOLD:
                prob_band = "Alta (≥90)"
            elif best_prob >= MID_HIGH:
                prob_band = "Média-Alta (80-89)"
            elif best_prob >= MID_LOW:
                prob_band = "Média-Baixa (60-79)"
            else:
                prob_band = "Baixa (<60)"

        results.append({
            "Company": company,
            "Classification": classification,
            "Best Model ACCT": best_acct,
            "Best Model ACCT NAME": best_acct_name,
            "Best Model Score": best_prob,
            "Best Model Level": best_level,
            "Expected ACCT(s)": ", ".join(sorted(exp_accts)) if exp_accts else "Não encontrado",
            "Score Band": prob_band,
            "Model Candidates": n_model_rows,
            "Multiple 100% Matches": has_multiple_100,
            "Detail": detail,
        })

    return pd.DataFrame(results)


# ── Métricas agregadas ────────────────────────────────────────────────────
def compute_metrics(classified_df):
    total = len(classified_df)
    tp = (classified_df["Classification"] == "True Positive").sum()
    tn = (classified_df["Classification"] == "True Negative").sum()
    fp = (classified_df["Classification"] == "False Positive").sum()
    fn = (classified_df["Classification"] == "False Negative").sum()
    partial = (classified_df["Classification"] == "Partial Match").sum()
    wrong = (classified_df["Classification"] == "Wrong Match").sum()

    correct = tp + tn
    accuracy = round(correct / total * 100, 2) if total else 0
    accuracy_lenient = round((correct + partial) / total * 100, 2) if total else 0

    precision_denom = tp + fp + partial + wrong
    precision = round(tp / precision_denom * 100, 2) if precision_denom else 0

    recall_denom = tp + fn + partial + wrong
    recall = round(tp / recall_denom * 100, 2) if recall_denom else 0

    f1 = round(2 * precision * recall / (precision + recall), 2) if (precision + recall) else 0

    tn_rate = round(tn / (tn + fp) * 100, 2) if (tn + fp) else 0

    multi_100 = classified_df["Multiple 100% Matches"].sum()

    band_counts = classified_df["Score Band"].value_counts().to_dict()

    fp_details = classified_df[classified_df["Classification"] == "False Positive"][
        ["Company", "Best Model Score", "Detail"]
    ].to_dict("records")

    rows = [
        ("Visão Geral", "Total de empresas avaliadas", total),
        ("Visão Geral", "Empresas com match correto (True Positive)", int(tp)),
        ("Visão Geral", "Empresas corretamente não encontradas (True Negative)", int(tn)),
        ("Visão Geral", "Falsos positivos (match indevido)", int(fp)),
        ("Visão Geral", "Falsos negativos (não encontrou, mas deveria)", int(fn)),
        ("Visão Geral", "Match parcial (esperado está entre candidatos, mas não é o top-1)", int(partial)),
        ("Visão Geral", "Match errado (recomendação incorreta)", int(wrong)),
        ("", "", ""),
        ("Acurácia", "Acurácia estrita (TP + TN) / Total", f"{accuracy}%"),
        ("Acurácia", "Acurácia leniente (TP + TN + Partial) / Total", f"{accuracy_lenient}%"),
        ("Acurácia", "Precisão (TP / matches feitos)", f"{precision}%"),
        ("Acurácia", "Recall (TP / matches que deveriam existir)", f"{recall}%"),
        ("Acurácia", "F1-Score", f"{f1}%"),
        ("Acurácia", "Taxa de True Negative (TN / (TN + FP))", f"{tn_rate}%"),
        ("", "", ""),
        ("Distribuição de Score", "Alta confiança (≥90)", band_counts.get("Alta (≥90)", 0)),
        ("Distribuição de Score", "Média-Alta (80-89)", band_counts.get("Média-Alta (80-89)", 0)),
        ("Distribuição de Score", "Média-Baixa (60-79)", band_counts.get("Média-Baixa (60-79)", 0)),
        ("Distribuição de Score", "Baixa (<60)", band_counts.get("Baixa (<60)", 0)),
        ("Distribuição de Score", "Sem score (não encontrado)", band_counts.get("", 0)),
        ("", "", ""),
        ("Qualidade de Dados", "Empresas com múltiplos matches 100% (possível lixo)", int(multi_100)),
        ("", "", ""),
        ("Análise de Falsos Positivos", "Detalhamento", ""),
    ]

    for fp_item in fp_details:
        rows.append((
            "Análise de Falsos Positivos",
            fp_item["Company"],
            f"Score {fp_item['Best Model Score']} - {fp_item['Detail']}"
        ))

    # Análise da faixa 60-80
    mid_range = classified_df[classified_df["Score Band"].isin(["Média-Baixa (60-79)"])]
    mid_correct = mid_range["Classification"].isin(["True Positive"]).sum()
    mid_wrong = mid_range["Classification"].isin(["False Positive", "Wrong Match"]).sum()
    mid_total = len(mid_range)

    rows.append(("", "", ""))
    rows.append(("Análise Faixa 60-79", "Total de empresas nesta faixa", int(mid_total)))
    rows.append(("Análise Faixa 60-79", "Matches corretos nesta faixa", int(mid_correct)))
    rows.append(("Análise Faixa 60-79", "Matches incorretos ou falsos positivos", int(mid_wrong)))
    if mid_total:
        rows.append(("Análise Faixa 60-79", "Taxa de erro nesta faixa", f"{round(mid_wrong/mid_total*100,2)}%"))
    rows.append(("Análise Faixa 60-79", "Recomendação",
        "Alta taxa de erro sugere elevar o MID_CONFIDENCE_THRESHOLD para ≥80."
        if mid_total and mid_wrong / mid_total > 0.5
        else "Taxa de erro aceitável para a faixa."
    ))

    return pd.DataFrame(rows, columns=["Section", "Metric", "Value"])


# ── Casos especiais para análise narrativa ────────────────────────────────
def build_special_cases(classified_df, model_df):
    cases = []

    # Caso AS AMERICA
    as_am = classified_df[classified_df["Company"].str.contains("As America", case=False, na=False)]
    if not as_am.empty:
        row = as_am.iloc[0]
        mod_rows = get_all_model_rows(model_df, row["Company"])
        candidates = ", ".join(
            f"{r['ACCT NAME']} (score {r['Probability']})"
            for _, r in mod_rows.iterrows()
        )
        cases.append({
            "Caso": "AS AMERICA INC — Cliente inexistente com falsos matches",
            "Descrição": (
                f"A empresa '{row['Company']}' não existia na base corporativa. "
                f"Apesar disso, o modelo retornou {len(mod_rows)} candidatos: {candidates}. "
                f"Nenhum desses matches se concretizou. O cliente foi posteriormente criado na base "
                f"após a análise, evidenciando que nomes genéricos podem gerar falsos positivos "
                f"com scores surpreendentemente altos."
            ),
        })

    # Casos com múltiplos 100%
    multi_100 = classified_df[classified_df["Multiple 100% Matches"]]
    for _, row in multi_100.iterrows():
        mod_rows = get_all_model_rows(model_df, row["Company"])
        rows_100 = mod_rows[mod_rows["Probability"] == 100]
        accts = ", ".join(f"{r['ACCT NAME']} ({r['ACCT']})" for _, r in rows_100.iterrows())
        cases.append({
            "Caso": f"{row['Company']} — Múltiplos matches 100%",
            "Descrição": (
                f"A empresa retornou {rows_100['ACCT'].nunique()} contas distintas com score 100%: {accts}. "
                f"Isso indica possível presença de registros desatualizados ou relações "
                f"corporativas (M&A) que o modelo não consegue resolver sozinho. "
                f"Requer revisão manual com o time de Customer."
            ),
        })

    # Falsos positivos na faixa 60-80
    fp_mid = classified_df[
        (classified_df["Classification"] == "False Positive") &
        (classified_df["Score Band"].isin(["Média-Baixa (60-79)", "Média-Alta (80-89)"]))
    ]
    if not fp_mid.empty:
        names = ", ".join(fp_mid["Company"].tolist())
        cases.append({
            "Caso": "Falsos positivos na faixa 60-80",
            "Descrição": (
                f"As empresas {names} receberam matches com scores entre 60 e 80, "
                f"mas o resultado correto era 'Não encontrado'. Essa concentração de erros "
                f"na faixa intermediária sugere que elevar o threshold mínimo de fallback "
                f"(MID_CONFIDENCE_THRESHOLD) para ≥80 reduziria falsos positivos sem impactar "
                f"significativamente a cobertura de matches legítimos."
            ),
        })

    return pd.DataFrame(cases)


# ── Geração do relatório Excel ────────────────────────────────────────────
def write_report(classified_df, metrics_df, special_cases_df, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metrics_df.to_excel(writer, sheet_name="Métricas", index=False)
        classified_df.to_excel(writer, sheet_name="Detalhamento", index=False)
        special_cases_df.to_excel(writer, sheet_name="Casos Especiais", index=False)

    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", start_color="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    color_map = {
        "True Positive": PatternFill("solid", start_color="C6EFCE"),
        "True Negative": PatternFill("solid", start_color="D9E2F3"),
        "False Positive": PatternFill("solid", start_color="FFC7CE"),
        "False Negative": PatternFill("solid", start_color="FCE4D6"),
        "Partial Match": PatternFill("solid", start_color="FFFFCC"),
        "Wrong Match": PatternFill("solid", start_color="F4B084"),
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    # Aplicar cores na aba Detalhamento
    ws_detail = wb["Detalhamento"]
    class_col = None
    for cell in ws_detail[1]:
        if cell.value == "Classification":
            class_col = cell.column - 1
            break

    if class_col is not None:
        for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row):
            classification = row[class_col].value
            fill = color_map.get(classification)
            if fill:
                for cell in row:
                    cell.fill = fill

    wb.save(output_path)


# ── Main ──────────────────────────────────────────────────────────────────
def main():
    print("Carregando dados...")
    model_df, expected_df = load_data(INPUT_FILE)

    print("Classificando empresas...")
    classified_df = classify_companies(model_df, expected_df)

    print("Calculando métricas...")
    metrics_df = compute_metrics(classified_df)

    print("Analisando casos especiais...")
    special_cases_df = build_special_cases(classified_df, model_df)

    print(f"Gerando relatório em {OUTPUT_FILE}...")
    write_report(classified_df, metrics_df, special_cases_df, OUTPUT_FILE)

    # Print resumo no terminal
    print("\n" + "=" * 60)
    print("RESUMO DA AVALIAÇÃO")
    print("=" * 60)
    for _, row in metrics_df.iterrows():
        if row["Section"] == "":
            print()
        else:
            print(f"  [{row['Section']}] {row['Metric']}: {row['Value']}")
    print("=" * 60)
    print(f"\nRelatório salvo em: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
