import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def _append_rows(summary_rows, df_rows):
    for _, row in df_rows.iterrows():
        row = row.copy()
        row["Probability"] = round(row["Probability"], 2)
        summary_rows.append(row.to_dict())


def _collapse_group_with_multiple_clients(summary_rows, df_rows):
    for acct_val, subdf in df_rows.groupby("ACCT"):
        if subdf["GBL_CLIENT_ID"].nunique(dropna=True) > 1:
            best_row = subdf.loc[subdf["Probability"].idxmax()].copy()
            best_row["GBL_CLIENT_ID"] = "Multiple Global Client IDs - check Details tab"
            best_row["GBL_CLIENT_NAME"] = ""
            best_row["Probability"] = round(best_row["Probability"], 2)
            summary_rows.append(best_row.to_dict())
        else:
            _append_rows(summary_rows, subdf)


def build_summary(df_details, df_input):
    summary_rows = []

    for company, group in df_details.groupby("Company"):
        group_12 = group[group["Group"].isin(["Group 1", "Group 2"])]
        high_12 = group_12[group_12["Probability"] >= 90]
        if not high_12.empty:
            _append_rows(summary_rows, high_12)
            continue

        group3 = group[group["Group"] == "Group 3"]
        high_3 = group3[group3["Probability"] >= 90]
        if not high_3.empty:
            non_short_3 = high_3[high_3["ACCT LEVEL"] == "GBL_BUY_GRP"]
            final_3 = non_short_3.copy() if not non_short_3.empty else high_3.copy()
            _collapse_group_with_multiple_clients(summary_rows, final_3)
            continue

        group4 = group[group["Group"] == "Group 4"]
        high_4 = group4[group4["Probability"] >= 90]
        if not high_4.empty:
            non_short_4 = high_4[high_4["ACCT LEVEL"] == "GBL_BUY_GRP"]
            final_4 = non_short_4.copy() if not non_short_4.empty else high_4.copy()
            _collapse_group_with_multiple_clients(summary_rows, final_4)
            continue

        mid_candidates = group[(group["Probability"] >= 50) & (group["Probability"] < 90)]
        if not mid_candidates.empty:
            best_row = mid_candidates.loc[mid_candidates["Probability"].idxmax()].copy()
            best_row["Probability"] = round(best_row["Probability"], 2)
            summary_rows.append(best_row.to_dict())
        else:
            summary_rows.append({
                "Company": company,
                "ACCT LEVEL": "Customer not found",
                "ACCT": None,
                "ACCT NAME": None,
                "Probability": None,
                "COV_TYPE_ID": None,
                "COV_NAME": None,
                "GBL_BUY_GRP": None,
                "GBL_BUY_GRP_NAME": None,
                "DOM_BUY_GRP": None,
                "DOM_BUY_GRP_NAME": None,
                "GBL_CLIENT_ID": None,
                "GBL_CLIENT_NAME": None,
                "acct_list_id": None,
                "acct_list_name": None,
                "Group": None,
                "MATCH_SOURCE": None,
                "CUST NAME": None,
                "Industry": None,
            })

    df_summary = pd.DataFrame(summary_rows)
    df_summary.sort_values(by="Company", inplace=True)

    subset_cols = [
        "Company", "ACCT LEVEL", "ACCT", "ACCT NAME", "COV_TYPE_ID", "COV_NAME",
        "GBL_BUY_GRP", "GBL_BUY_GRP_NAME", "DOM_BUY_GRP", "DOM_BUY_GRP_NAME",
        "GBL_CLIENT_ID", "GBL_CLIENT_NAME", "acct_list_id", "acct_list_name"
    ]
    actual_subset_cols = [col for col in subset_cols if col in df_summary.columns]
    df_summary.drop_duplicates(subset=actual_subset_cols, inplace=True)

    if "Group" in df_summary.columns:
        df_summary.drop(columns=["Group"], inplace=True)

    return df_summary

def write_summary_excel(df_summary, df_details, output_file, metrics_df=None):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_details.to_excel(writer, sheet_name="Details", index=False)
        if metrics_df is not None and not metrics_df.empty:
            metrics_df.to_excel(writer, sheet_name="Metrics", index=False)

    wb = openpyxl.load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 1

        light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.fill = light_grey_fill

    summary_ws = wb["Summary"]

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Descobrir índices das colunas pelo nome do header
    header_map = {}
    for cell in summary_ws[1]:
        header_map[cell.value] = cell.column

    company_col = header_map.get("Company")
    acct_level_col = header_map.get("ACCT LEVEL")
    acct_list_name_col = header_map.get("acct_list_name")

    company_counts = {}
    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
        company_val = row[company_col - 1].value if company_col else None
        if company_val:
            company_counts[company_val] = company_counts.get(company_val, 0) + 1

    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
        company_val = row[company_col - 1].value if company_col else None
        acct_level_val = row[acct_level_col - 1].value if acct_level_col else None
        acct_list_name_val = row[acct_list_name_col - 1].value if acct_list_name_col else None

        if acct_level_val == "Customer not found":
            for cell in row:
                cell.fill = yellow_fill

        elif acct_list_name_val == "Activate Unassigned":
            for cell in row:
                cell.fill = light_orange_fill

        elif company_counts.get(company_val, 0) > 1:
            for cell in row:
                cell.fill = light_blue_fill

    wb.save(output_file)